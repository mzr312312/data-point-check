import tkinter as tk
from tkinter import filedialog, scrolledtext
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import copy


# ========== 1. 解析 markdown 字典 ==========
def parse_markdown_dict(md_file):
    dictionary = {}
    current_header = None

    with open(md_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line.startswith("##"):
                match = re.match(r'##\s*(.+)', line)
                if match:
                    header = match.group(1).strip().replace('\\n', '\n')
                    current_header = header
                    dictionary[current_header] = []
            elif line.startswith("-") and current_header:
                item = line[1:].strip()
                if item == "（此列为必填，但无固定枚举值）":
                    dictionary[current_header] = []  # 空列表表示无需枚举校验
                else:
                    dictionary[current_header].append(item)

    return dictionary


# ========== 2. 单元格校验函数 ==========
def validate_cell(value, column_name, dictionary):
    value = str(value).strip() if value is not None else ""

    # 清理全角空格、换行符、制表符
    value = re.sub(r'[\u3000\n\r\t]', '', value)

    # 必填项检查
    if value == "":
        return "为空", value

    # 枚举值检查（如果该列有枚举值）
    enum_values = dictionary.get(column_name, [])
    if len(enum_values) > 0 and value not in enum_values:
        return "与字典不符", value

    return "通过", value  # 默认通过


# ========== 4. 主程序类 ==========
class ExcelValidatorApp:
    def __init__(self, root, dictionary):
        self.root = root
        self.dictionary = dictionary
        self.cell_errors = []  # 存储单元格校验错误 (row_idx, col_name)
        self.group_errors = []  # 存储分组一致性错误 (device_name, row_idx, col_name, ref_value)

        # 创建 GUI
        self.create_gui()

    def create_gui(self):
        self.root.title("Excel采集点校验工具")

        # 按钮区域
        self.btn_frame = tk.Frame(self.root)
        self.btn_frame.pack(pady=10)

        self.select_button = tk.Button(self.btn_frame, text="选择Excel文件", command=self.load_excel)
        self.select_button.pack(side=tk.LEFT, padx=5)

        # 输出区域（日志显示）
        self.output = scrolledtext.ScrolledText(self.root, width=100, height=30, wrap=tk.WORD)
        self.output.pack(padx=10, pady=10)

    def log_message(self, msg):
        """向 GUI 输出日志信息，删除所有换行符"""
        # 替换所有换行符和全角空格
        msg = msg.replace('\n', ' ').replace('\r', ' ')
        self.output.insert(tk.END, msg + "\n")
        self.output.see(tk.END)

    def load_excel(self):
        self.cell_errors = []  # 清空单元格错误记录
        self.group_errors = []  # 清空分组错误记录
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        self.log_message(f"正在加载文件：{file_path}")
        try:
            # 提示文件加载进度
            self.log_message("⏳ 开始解析 Excel 文件...")
            df = pd.read_excel(file_path, sheet_name='采集点', engine='openpyxl', header=1)
            headers = [str(col).strip() for col in df.columns]
            stats = {col: {"total": 0, "pass": 0, "fail": 0} for col in headers if col in self.dictionary}
            error_count = 0

            # 提示逐行校验进度
            total_rows = len(df)
            self.log_message(f"⏳ 开始逐行校验，共 {total_rows} 行数据...")
            for row_idx, row in df.iterrows():
                actual_row_number = row_idx + 3  # 修正：pandas索引 + 3（从第3行开始）
                # # 每隔 100 行打印一次进度
                # if row_idx % 100 == 0:
                #     self.log_message(f"⏳ 正在校验第 {row_idx + 1} 行（共 {total_rows} 行）...")

                row_has_content = any(str(val).strip() != '' for val in row.values)
                if not row_has_content:
                    continue

                for col_name in headers:
                    if col_name not in self.dictionary:
                        continue
                    cell_value = row[col_name]
                    result, cleaned_value = validate_cell(cell_value, col_name, self.dictionary)
                    stats[col_name]["total"] += 1

                    display_col_name = col_name.replace('\n', ' ')
                    if result == "为空":
                        msg = f"【空值】'行号'{actual_row_number}'列名'{display_col_name}）'为空'"
                        self.log_message(msg)
                        error_count += 1
                        stats[col_name]["fail"] += 1
                        self.cell_errors.append((row_idx, col_name))
                    elif result == "与字典不符":
                        msg = f"【与字典不符】'行号'{actual_row_number}'列名'{display_col_name}）'原始值'{cell_value}'清理后值'{cleaned_value}'与字典不符'"
                        self.log_message(msg)
                        error_count += 1
                        stats[col_name]["fail"] += 1
                        self.cell_errors.append((row_idx, col_name))
                    else:
                        stats[col_name]["pass"] += 1

            # 输出每列统计信息
            self.log_message("\n📊 每列校验结果统计：")
            for col_name, stat in stats.items():
                if stat["total"] > 0:
                    display_col_name = col_name.replace('\n', ' ')
                    self.log_message(
                        f"{display_col_name} 共检查 {stat['total']} 项，通过 {stat['pass']} 项，失败 {stat['fail']} 项"
                    )

            # 分组一致性校验
            self.validate_group_consistency(df, stats, headers, error_count)

            # 最终结果提示
            if error_count == 0:
                self.log_message("✅ 校验完成！\n")
            else:
                self.log_message(f"⚠️ 校验完成，发现 {error_count} 个问题！\n")
                # 生成错误标黄文件
                # self.save_error_files(file_path, headers, df)

        except Exception as e:
            self.log_message(f"⚠️ 发生异常：{str(e)}")

    def validate_group_consistency(self, df, stats, headers, error_count):
        """
        针对同一设备名称的采集点，检查指定字段是否一致
        使用 value_counts 找出出现次数最多的值作为参考值，避免 mode() 返回多个值的问题
        """
        group_by_column = "设备名称\n（必填）"
        check_columns = [
            "基地\n（必选）",
            "车间\n（必选）",
            "工段\n（必选）",
            "工序/系统\n（必选）",
            "设备子类型\n（必选）"
        ]
        valid_check_columns = [col for col in check_columns if col in headers]
        grouped = df.groupby(group_by_column)
        self.log_message("\n🔍 开始校验同一设备名称下的字段一致性...")

        total_groups = len(grouped)
        for i, (device_name, group) in enumerate(grouped):
            # 每隔 10 组打印一次进度
            # if i % 10 == 0:
            #     self.log_message(f"⏳ 正在校验第 {i + 1} 组设备（共 {total_groups} 组）：设备名称 '{device_name}'...")

            if len(group) <= 1:
                continue

            for col in valid_check_columns:
                values = group[col].astype(str).str.strip()
                unique_values = values.unique()
                if len(unique_values) > 1:
                    value_counts = values.value_counts()
                    mode_value = value_counts.idxmax()
                    for idx, row in group.iterrows():
                        actual_row_number = idx + 3
                        value = str(row[col]).strip()
                        if value != mode_value:
                            msg = f"【设备名称校验】'设备名称'{device_name}' 列名 '{col}' 行号' {actual_row_number} '当前值'{value}'参考值'{mode_value}'"
                            self.log_message(msg)
                            error_count += 1
                            stats[col]["fail"] += 1
                            self.group_errors.append((device_name, idx, col, mode_value))

        self.log_message("✅ 分组一致性校验完成！\n")

    def save_error_files(self, original_path, headers, df):
        """生成两个文件：报错文件和自动修改文件"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            wb = load_workbook(original_path)
            ws = wb['采集点']
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            auto_modified_df = df.copy()

            # 提示文件保存进度
            self.log_message("⏳ 正在生成报错文件...")
            error_wb = load_workbook(original_path)
            error_ws = error_wb['采集点']

            for row_idx, col_name in self.cell_errors:
                if col_name in headers:
                    col_index = headers.index(col_name)
                    excel_row = row_idx + 3
                    col_letter = get_column_letter(col_index + 1)
                    cell = error_ws[f"{col_letter}{excel_row}"]
                    cell.fill = yellow_fill

            for device_name, row_idx, col_name, ref_value in self.group_errors:
                if col_name in headers:
                    col_index = headers.index(col_name)
                    excel_row = row_idx + 3
                    col_letter = get_column_letter(col_index + 1)
                    cell = error_ws[f"{col_letter}{excel_row}"]
                    cell.fill = yellow_fill

            error_file = f"报错文件_{timestamp}.xlsx"
            # error_wb.save(error_file)
            self.log_message(f"📄 已生成报错文件：{error_file}")

            # 提示自动修改文件生成进度
            self.log_message("⏳ 正在生成自动修改文件...")
            auto_wb = load_workbook(original_path)
            auto_ws = auto_wb['采集点']

            for row_idx, col_name in self.cell_errors:
                if col_name in headers:
                    col_index = headers.index(col_name)
                    excel_row = row_idx + 3
                    col_letter = get_column_letter(col_index + 1)
                    cell = auto_ws[f"{col_letter}{excel_row}"]
                    cell.fill = yellow_fill

            for device_name, row_idx, col_name, ref_value in self.group_errors:
                if col_name in headers:
                    col_index = headers.index(col_name)
                    excel_row = row_idx + 3
                    col_letter = get_column_letter(col_index + 1)
                    cell = auto_ws[f"{col_letter}{excel_row}"]
                    cell.fill = yellow_fill
                    cell.value = ref_value
                    auto_modified_df.at[row_idx, col_name] = ref_value

            # auto_file = f"自动修改文件_{timestamp}.xlsx"
            # auto_wb.save(auto_file)
            self.log_message(f"📄 已生成自动修改文件：{auto_file}")

        except Exception as e:
            self.log_message(f"⚠️ ：{str(e)}")


# ========== 5. 启动程序 ==========
if __name__ == "__main__":
    md_file = "采集表校验字典.md"
    dictionary = parse_markdown_dict(md_file)

    root = tk.Tk()
    app = ExcelValidatorApp(root, dictionary)
    root.mainloop()